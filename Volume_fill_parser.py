import streamlit as st
import os
import PyPDF2
import openpyxl
import re
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from io import BytesIO
from openpyxl.drawing.image import Image
from typing import List, Tuple, Dict
from dataclasses import dataclass
import tempfile

@dataclass
class ProcessingConfig:
    """Configuration settings for data processing"""
    lower_cutoff: float
    upper_cutoff: float

class DataProcessor:
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.workbook = openpyxl.Workbook()
        
    def extract_pdf_text(self, file) -> str:
        """Extract text from PDF file starting from page 4"""
        reader = PyPDF2.PdfReader(file)
        text = ""
        for page_num in range(3, len(reader.pages)):
            page_text = reader.pages[page_num].extract_text()
            text += page_text.replace('\xa0', ' ')
        return text

    def parse_data_lines(self, text: str) -> Tuple[List[str], List[str]]:
        """Parse text into header and data lines"""
        lines = text.split('\n')
        header_line = None
        data_lines = []
        
        for line in lines:
            # Skip lines containing page numbers or specific headers
            if re.search(r'[Ss]eite\s+\d+', line) or re.search(r'FlexLine\s+M60\s+Nettogewichte\s+je\s+Charge', line):
                continue
            
            if re.search(r'Pos\.\s+0[1-4]', line) and not header_line:
                header_line = line
            elif not re.search(r'Pos\.\s+0[1-4]', line):
                data_lines.append(line)
                
        return header_line, data_lines

    def _write_headers(self, worksheet):
        """Write headers to worksheet"""
        custom_headers = [
            "Pos. 01", "Unit", 
            "Pos. 02", "Unit", 
            "Pos. 03", "Unit", 
            "Pos. 04", "Unit", 
            "Datum/Uhrzeit"
        ]
        
        for col_num, header in enumerate(custom_headers, start=1):
            worksheet.cell(row=1, column=col_num, value=header)

    def process_data_lines(self, data_lines: List[str], worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[List[List[float]], List[str], List[List[float]]]:
        """Process data lines and write to worksheet"""
        pos_values = [[] for _ in range(4)]
        pos_units = ["" for _ in range(4)]
        all_pos_values = [[] for _ in range(4)]
        
        for row_num, line in enumerate(data_lines, start=2):
            tokens = line.split()
            if len(tokens) < 8:
                continue
                
            self._process_line_tokens(tokens, row_num, worksheet, pos_values, pos_units, all_pos_values)
            
        return pos_values, pos_units, all_pos_values

    def _process_line_tokens(self, tokens: List[str], row_num: int, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           pos_values: List[List[float]], pos_units: List[str], all_pos_values: List[List[float]]):
        """Process individual line tokens"""
        col_index = 1
        for i in range(0, 8, 2):
            if i+1 < len(tokens):
                pos_index = i // 2
                try:
                    value = float(tokens[i])
                    if value != 0:
                        all_pos_values[pos_index].append(value)
                        pos_units[pos_index] = tokens[i+1]
                    
                    if value != 0 and self.config.lower_cutoff <= value <= self.config.upper_cutoff:
                        pos_values[pos_index].append(value)
                except ValueError:
                    pass
                
                worksheet.cell(row=row_num, column=col_index, value=tokens[i])
                worksheet.cell(row=row_num, column=col_index + 1, value=tokens[i+1])
                col_index += 2
        
        if len(tokens) >= 10:
            worksheet.cell(row=row_num, column=col_index, value=f"{tokens[8]} {tokens[9]}")

    def _calculate_statistics(self, values):
        """Calculate statistics for the values"""
        return {
            'mean': np.mean(values),
            'std_dev': np.std(values),
            'min': np.min(values),
            'max': np.max(values)
        }

    def _plot_histogram_and_curve(self, values, stats_data, unit):
        """Plot histogram and normal distribution curve"""
        n, bins, patches = plt.hist(values, bins=15, density=True, alpha=0.8, color='#D9E6F2')
        
        x = np.linspace(stats_data['min'] - stats_data['std_dev'], 
                        stats_data['max'] + stats_data['std_dev'], 100)
        y = stats.norm.pdf(x, stats_data['mean'], stats_data['std_dev'])
        plt.plot(x, y, color='#4472C4', linewidth=2.5)

    def _add_vertical_lines(self, stats_data, unit):
        """Add vertical lines for statistics and cutoffs"""
        plt.axvline(x=stats_data['min'], color='#9370DB', linestyle='--', linewidth=1.5, 
                   label=f'Min: {stats_data["min"]:.4f} {unit}')
        plt.axvline(x=stats_data['max'], color='#90EE90', linestyle='--', linewidth=1.5, 
                   label=f'Max: {stats_data["max"]:.4f} {unit}')
        plt.axvline(x=stats_data['mean'], color='#FFB6C1', linestyle='-', linewidth=1.5, 
                   label=f'Mean: {stats_data["mean"]:.4f} {unit}')
        
        plt.axvline(x=self.config.lower_cutoff, color='#FF8C00', linestyle=':', linewidth=1.5,
                   label=f'Lower cutoff: {self.config.lower_cutoff:.4f} {unit}')
        plt.axvline(x=self.config.upper_cutoff, color='#FF8C00', linestyle=':', linewidth=1.5,
                   label=f'Upper cutoff: {self.config.upper_cutoff:.4f} {unit}')

    def _set_plot_attributes(self, pos_num, stats_data, unit):
        """Set plot title, labels, and other attributes"""
        plt.title(f'Distribution for Pos. 0{pos_num} (Std Dev: {stats_data["std_dev"]:.4f} {unit})', fontsize=14)
        plt.xlabel(f'Value ({unit})', fontsize=12)
        plt.ylabel('Density', fontsize=12)
        plt.legend(fontsize=10)
        plt.grid(True, alpha=0.2)
        plt.tight_layout()

    def _add_graph_to_sheet(self, img_data, pos_num, stats_data, unit, graph_sheet, values_count, all_values_count):
        """Add graph and statistics to the graph sheet"""
        img = Image(img_data)
        graph_sheet.add_image(img, f'A{(pos_num-1) * 30 + 1}')
        
        row_offset = (pos_num-1) * 30 + 22
        graph_sheet.cell(row=row_offset, column=1, value=f"Statistics for Pos. 0{pos_num}")
        graph_sheet.cell(row=row_offset+1, column=1, value=f"Mean: {stats_data['mean']:.4f} {unit}")
        graph_sheet.cell(row=row_offset+2, column=1, value=f"Std Dev: {stats_data['std_dev']:.4f} {unit}")
        graph_sheet.cell(row=row_offset+3, column=1, value=f"Min: {stats_data['min']:.4f} {unit}")
        graph_sheet.cell(row=row_offset+4, column=1, value=f"Max: {stats_data['max']:.4f} {unit}")
        graph_sheet.cell(row=row_offset+5, column=1, value=f"Count (within cutoffs): {values_count}")
        graph_sheet.cell(row=row_offset+6, column=1, value=f"Total non-zero values: {all_values_count}")
        graph_sheet.cell(row=row_offset+7, column=1, value=f"Cutoff range: {self.config.lower_cutoff:.4f} - {self.config.upper_cutoff:.4f} {unit}")

    def create_distribution_graph(self, values: List[float], pos_num: int, unit: str, 
                                graph_sheet: openpyxl.worksheet.worksheet.Worksheet, all_values_count: int) -> None:
        """Create and save distribution graph"""
        if not values:
            return
            
        stats_data = self._calculate_statistics(values)
        
        plt.figure(figsize=(10, 6), facecolor='white')
        ax = plt.gca()
        ax.set_facecolor('white')
        
        self._plot_histogram_and_curve(values, stats_data, unit)
        self._add_vertical_lines(stats_data, unit)
        self._set_plot_attributes(pos_num, stats_data, unit)
        
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=100, bbox_inches='tight')
        img_data.seek(0)
        plt.close()
        
        self._add_graph_to_sheet(img_data, pos_num, stats_data, unit, graph_sheet, len(values), all_values_count)

    def process_file(self, file, filename: str) -> None:
        """Process individual PDF file"""
        text = self.extract_pdf_text(file)
        
        sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet names limited to 31 chars
        worksheet = self.workbook.create_sheet(title=sheet_name)
        graph_sheet = self.workbook.create_sheet(title=f"{sheet_name[:27]}_Graphs")
        
        self._write_headers(worksheet)
        
        _, data_lines = self.parse_data_lines(text)
        pos_values, pos_units, all_pos_values = self.process_data_lines(data_lines, worksheet)
        
        for pos_index, values in enumerate(pos_values):
            self.create_distribution_graph(
                values, 
                pos_index + 1, 
                pos_units[pos_index], 
                graph_sheet, 
                len(all_pos_values[pos_index])
            )

    def process_files(self, files) -> BytesIO:
        """Process multiple PDF files and return Excel workbook"""
        for file in files:
            self.process_file(file, file.name)
        
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
            
        excel_data = BytesIO()
        self.workbook.save(excel_data)
        excel_data.seek(0)
        
        return excel_data

def main():
    st.set_page_config(page_title="Volume Fill Data Parser", layout="wide")
    
    st.title("Volume Fill Data Parser")
    st.markdown("""
    This application processes PDF files containing measurement data, extracts values, 
    performs statistical analysis, and generates an Excel file with data and distribution graphs.
    """)
    
    with st.sidebar:
        st.header("Settings")
        lower_cutoff = st.number_input("Lower Cutoff Value", value=1.0, step=0.1, format="%.1f")
        upper_cutoff = st.number_input("Upper Cutoff Value", value=3.0, step=0.1, format="%.1f")
        
        # Add filename input in sidebar
        output_filename = st.text_input(
            "Output Filename",
            value="pdf_contents_with_graphs",
            help="Enter the desired filename (without .xlsx extension)"
        )
        
        st.info("""
        **Note:** Statistical analysis will only include values between the lower and upper cutoffs.
        All non-zero values will still be displayed in the Excel file.
        """)
    
    uploaded_files = st.file_uploader("Upload PDF Files", type="pdf", accept_multiple_files=True)
    
    if uploaded_files:
        st.write(f"📁 Number of files uploaded: **{len(uploaded_files)}**")
        
        with st.expander("View uploaded files"):
            for file in uploaded_files:
                st.write(f"- {file.name}")
        
        process_button = st.button("Process Files")
        
        if process_button:
            # Validate filename
            if not output_filename.strip():
                st.error("Please enter a valid filename")
                return
                
            with st.spinner("Processing files... This may take a moment."):
                try:
                    config = ProcessingConfig(
                        lower_cutoff=lower_cutoff,
                        upper_cutoff=upper_cutoff
                    )
                    
                    processor = DataProcessor(config)
                    
                    # Create progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Process each file with progress updates
                    for idx, file in enumerate(uploaded_files):
                        status_text.text(f"Processing {file.name}...")
                        file.seek(0)  # Reset file pointer
                        progress = (idx + 1) / len(uploaded_files)
                        progress_bar.progress(progress)
                    
                    # Process all files
                    excel_data = processor.process_files(uploaded_files)
                    
                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()
                    
                    st.success("✅ Processing complete!")
                    
                    # Sanitize filename
                    safe_filename = re.sub(r'[<>:"/\\|?*]', '', output_filename.strip())
                    
                    # Provide download button with custom filename
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_data,
                        file_name=f"{safe_filename}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"An error occurred during processing: {str(e)}")
    else:
        st.info("👆 Please upload one or more PDF files to begin")

if __name__ == "__main__":
    main()